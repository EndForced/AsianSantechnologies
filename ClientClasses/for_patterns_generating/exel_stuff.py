# from PIL.ImagePalette import random
from openpyxl import load_workbook, Workbook
import openpyxl
import random

max_str_num = 22
#check pats_all exel file if interested

def write_matrices_to_excel(filename, sheet_name, matrices, append=True):
    """Writes matrices to an existing Excel file, one matrix per cell.

    Args:
        filename: Path to the Excel file.
        sheet_name: Name of the sheet to write to. Creates if it doesn't exist.
        matrices: A list of matrices (lists of lists).
        append: Boolean; if True, appends to the end of the sheet. If False, overwrites the sheet. (default: True)
    """
    try:
        workbook = load_workbook(filename)
        sheet = workbook[sheet_name] if sheet_name in workbook else workbook.create_sheet(sheet_name)

        if not append:
            sheet.delete_rows(1, sheet.max_row)

        next_row = sheet.max_row + 1 if append else 1

        for matrix in matrices:
            matrix_str = str(matrix)  # Convert matrix to string representation

            # Handle potential large matrix strings: Truncate if necessary
            max_cell_length = 32767 # Approximate limit for Excel cell text
            matrix_str = matrix_str[:max_cell_length] # Truncate if it exceeds Excel limit


            cell = sheet.cell(row=next_row, column=1) # Write to column A
            cell.value = matrix_str
            next_row +=1

        workbook.save(filename)
        print(f"Matrices written to '{filename}' in sheet '{sheet_name}'.")

    except FileNotFoundError:
        print(f"Error: File '{filename}' not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def write(data):
    write_matrices_to_excel("pats.xlsx", "pats_all", data)

def write_to_existing():
    transfer_and_clear_excel("pats.xlsx","pats_all.xlsx")

def create_file():
    workbook = Workbook()
    sheet1 = workbook.active
    sheet1.title = "pats_all"  # Rename the default sheet

    sheet2 = workbook.create_sheet("pats_all")  # Create a new sheet named "Sheet2"

    # Add data to sheet1

    workbook.save("pats_all.xlsx")
    print("file created")

def transfer_and_clear_excel(source_file, destination_file, sheet_name="pats_all"):
    """
    Transfers data from a source Excel file to a destination file, then clears the source file.

    Args:
        source_file: Path to the source Excel file.
        destination_file: Path to the destination Excel file.
        sheet_name: Name of the sheet to transfer data from.  (default="Sheet1")
    """

    try:
        # Load Workbooks
        source_wb = openpyxl.load_workbook(source_file)
        dest_wb = openpyxl.load_workbook(destination_file) if destination_file and openpyxl.workbook.Workbook != type(destination_file) else openpyxl.Workbook() #Open if there is an existing file, otherwise create it.


        # Select Sheets
        source_sheet = source_wb[sheet_name] if sheet_name in source_wb.sheetnames else None
        dest_sheet = dest_wb[sheet_name] if sheet_name in dest_wb.sheetnames else dest_wb.create_sheet(sheet_name)

        if source_sheet is None:
            print(f"Sheet '{sheet_name}' not found in source file.")
            return # Do not try to copy if source sheet doesnt exist

        # Transfer data
        max_row_src = source_sheet.max_row
        max_col_src = source_sheet.max_column

        next_row_dest = dest_sheet.max_row + 1 # Start appending in next empty row

        for row_num in range(1, max_row_src + 1):
            row_data = []
            for col_num in range (1, max_col_src+1):
              cell = source_sheet.cell(row=row_num, column=col_num)
              row_data.append(cell.value) #Get the value of cell

            for i, value in enumerate(row_data):
                dest_sheet.cell(row=next_row_dest, column=i+1).value = value #Set the value in the same location.
            next_row_dest += 1


        # Save the destination file
        dest_wb.save(destination_file)


        #Clear source file
        source_sheet.delete_rows(1, max_row_src)
        source_wb.save(source_file)


        print(f"Data transferred to '{destination_file}' and source file '{source_file}' cleared.")


    except FileNotFoundError:
        print(f"Error: One of the specified files not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

def rand_pat_from_file():
    global max_str_num
    workbook = load_workbook("pats_all.xlsx")
    sheet = workbook["pats_all"]
    pat = []
    while not pat:
        r = random.randint(2,max_str_num)
        pat = sheet.cell(row = r, column= 1).value
    return(pat)

def replace_ints_in_matrix_rev(matrix):
    replacements = {
        70: 10,
        10: 20,
        33: 32,
        30: 33,
        32: 34,
        7040: 41,
        7041: 42,
        1140: 51,
        1141: 52,
        60: 61,
        63: 62,
        61: 63,
        62: 64,
        7050: 71,
        7052: 72,
        7051: 73,
        7053: 74
    }

    new_matrix = []
    for row in matrix:
        new_row = []
        for element in row:
            if element in replacements:
                new_row.append(replacements[element])
            else:
                new_row.append(element)
        new_matrix.append(new_row)

    return new_matrix

create_file () # if pats all is corrupted, del it and run this def

